"""
gy数据处理
"""
from openpyxl import Workbook


def get_data(file):
    '''
    抽取数据
    :return: 数据时间，5组有效数据
    '''
    # 字段声明
    data_time = []
    lx = []
    ly = []  # 上归中
    rx = []
    ry = []
    extra = []
    with open(file) as file:
        data = file.readlines()
        if data[0] == '\n':
            data.remove(data[0])
        for i in data:
            data_item = i.strip()  # 去掉换行符
            data_item_list = data_item.split(',')  # 行数据用逗号分割成list
            data_time.append((data_item_list[0]))
            lx.append(data_item_list[1])  # lx列表
            ly.append(data_item_list[2])
            rx.append(data_item_list[3])
            ry.append(data_item_list[4])
            extra.append(data_item_list[5])
    return data_time, lx, ly, rx, ry, extra


def save_data(zuobiao, adc, yaogan, guizhong):
    '''
    保存数据
    :return: 数据保存excel，格式为xlsx
    '''
    data_time0, lx0, ly0, rx0, ry0, extra0 = get_data(file=f'56#/0万次/{adc}/{yaogan}/{yaogan}{guizhong}')
    data_time1, lx1, ly1, rx1, ry1, extra1 = get_data(file=f'56#/1万次/{adc}/{yaogan}/{yaogan}1万次{guizhong}')
    data_time2, lx2, ly2, rx2, ry2, extra2 = get_data(file=f'56#/2万次/{adc}/{yaogan}/{yaogan}2万次{guizhong}')
    data_time3, lx3, ly3, rx3, ry3, extra3 = get_data(file=f'56#/3万次/{adc}/{yaogan}/{yaogan}3万次{guizhong}')
    data_time4, lx4, ly4, rx4, ry4, extra4 = get_data(file=f'56#/4万次/{adc}/{yaogan}/{yaogan}4万次{guizhong}')
    data_time5, lx5, ly5, rx5, ry5, extra5 = get_data(file=f'56#/5万次/{adc}/{yaogan}/{yaogan}5万次{guizhong}')
    data_time6, lx6, ly6, rx6, ry6, extra6 = get_data(file=f'56#/6万次/{adc}/{yaogan}/{yaogan}6万次{guizhong}')
    data_time7, lx7, ly7, rx7, ry7, extra7 = get_data(file=f'56#/7万次/{adc}/{yaogan}/{yaogan}7万次{guizhong}')

    if zuobiao == 'LY':
        wb = Workbook()
        sheet = wb.active
        sheet.title = '摇杆数据分析'
        fields = [f'0次{yaogan}{guizhong}值', f'RY1万次{yaogan}{guizhong}值', f'RY2万次{yaogan}{guizhong}值',
                  f'RY3万次{yaogan}{guizhong}值', f'RY4万次{yaogan}上归中值', 'RY5万次左摇杆上归中值',
                  f'RY6万次{yaogan}{guizhong}值', f'RY7万次{yaogan}{guizhong}值']
        for i in range(len(fields)):  # 首行
            sheet.cell(row=1, column=i + 1).value = fields[i]
        for a, element in enumerate(ly0):  # 0次
            sheet.cell(row=a + 2, column=1).value = int(element)
        for b, element in enumerate(ly1):  # 1万次
            sheet.cell(row=b + 2, column=2).value = int(element)
        for i, element in enumerate(ly2):  # 2万次
            sheet.cell(row=i + 2, column=3).value = int(element)
        for i, element in enumerate(ly3):  # 3万次
            sheet.cell(row=i + 2, column=4).value = int(element)
        for i, element in enumerate(ly4):  # 4万次
            sheet.cell(row=i + 2, column=5).value = int(element)
        for i, element in enumerate(ly5):  # 5万次
            sheet.cell(row=i + 2, column=6).value = int(element)
        for i, element in enumerate(ly6):  # 6万次
            sheet.cell(row=i + 2, column=7).value = int(element)
        for i, element in enumerate(ly7):  # 7万次
            sheet.cell(row=i + 2, column=8).value = int(element)
        wb.save(f'{zuobiao}7万次{yaogan}{guizhong}值.xlsx')

    if zuobiao == 'LX':
        wb = Workbook()
        sheet = wb.active
        sheet.title = '摇杆数据分析'
        fields = [f'{zuobiao}0万次{yaogan}{guizhong}值', f'{zuobiao}1万次{yaogan}{guizhong}值',
                  f'{zuobiao}2万次{yaogan}{guizhong}值', f'{zuobiao}3万次{yaogan}{guizhong}值',
                  f'{zuobiao}4万次{yaogan}{guizhong}值', f'{zuobiao}5万次{yaogan}{guizhong}值',
                  f'{zuobiao}6万次{yaogan}{guizhong}值', f'{zuobiao}7万次{yaogan}{guizhong}值']
        for i in range(len(fields)):  # 首行
            sheet.cell(row=1, column=i + 1).value = fields[i]
        for a, element in enumerate(lx0):  # 0次
            sheet.cell(row=a + 2, column=1).value = int(element)
        for b, element in enumerate(lx1):  # 1万次
            sheet.cell(row=b + 2, column=2).value = int(element)
        for i, element in enumerate(lx2):  # 2万次
            sheet.cell(row=i + 2, column=3).value = int(element)
        for i, element in enumerate(lx3):  # 3万次
            sheet.cell(row=i + 2, column=4).value = int(element)
        for i, element in enumerate(lx4):  # 4万次
            sheet.cell(row=i + 2, column=5).value = int(element)
        for i, element in enumerate(lx5):  # 5万次
            sheet.cell(row=i + 2, column=6).value = int(element)
        for i, element in enumerate(lx6):  # 6万次
            sheet.cell(row=i + 2, column=7).value = int(element)
        for i, element in enumerate(lx7):  # 7万次
            sheet.cell(row=i + 2, column=8).value = int(element)
        wb.save(f'{zuobiao}7万次{yaogan}{guizhong}值.xlsx')

    if zuobiao == 'RX':
        wb = Workbook()
        sheet = wb.active
        sheet.title = '摇杆数据分析'
        fields = [f'{zuobiao}0万次{yaogan}{guizhong}值', f'{zuobiao}1万次{yaogan}{guizhong}值',
                  f'{zuobiao}2万次{yaogan}{guizhong}值', f'{zuobiao}3万次{yaogan}{guizhong}值',
                  f'{zuobiao}4万次{yaogan}{guizhong}值', f'{zuobiao}5万次{yaogan}{guizhong}值',
                  f'{zuobiao}6万次{yaogan}{guizhong}值', f'{zuobiao}7万次{yaogan}{guizhong}值']
        for i in range(len(fields)):  # 首行
            sheet.cell(row=1, column=i + 1).value = fields[i]
        for a, element in enumerate(rx0):  # 0次
            sheet.cell(row=a + 2, column=1).value = int(element)
        for b, element in enumerate(rx1):  # 1万次
            sheet.cell(row=b + 2, column=2).value = int(element)
        for i, element in enumerate(rx2):  # 2万次
            sheet.cell(row=i + 2, column=3).value = int(element)
        for i, element in enumerate(rx3):  # 3万次
            sheet.cell(row=i + 2, column=4).value = int(element)
        for i, element in enumerate(rx4):  # 4万次
            sheet.cell(row=i + 2, column=5).value = int(element)
        for i, element in enumerate(rx5):  # 5万次
            sheet.cell(row=i + 2, column=6).value = int(element)
        for i, element in enumerate(rx6):  # 6万次
            sheet.cell(row=i + 2, column=7).value = int(element)
        for i, element in enumerate(rx7):  # 7万次
            sheet.cell(row=i + 2, column=8).value = int(element)
        wb.save(f'{zuobiao}7万次{yaogan}{guizhong}值.xlsx')

    if zuobiao == 'RY':
        wb = Workbook()
        sheet = wb.active
        sheet.title = '摇杆数据分析'
        fields = [f'{zuobiao}0万次{yaogan}{guizhong}值', f'{zuobiao}1万次{yaogan}{guizhong}值',
                  f'{zuobiao}2万次{yaogan}{guizhong}值', f'{zuobiao}3万次{yaogan}{guizhong}值',
                  f'{zuobiao}4万次{yaogan}{guizhong}值', f'{zuobiao}5万次{yaogan}{guizhong}值',
                  f'{zuobiao}6万次{yaogan}{guizhong}值', f'{zuobiao}7万次{yaogan}{guizhong}值']
        for i in range(len(fields)):  # 首行
            sheet.cell(row=1, column=i + 1).value = fields[i]
        for a, element in enumerate(ry0):  # 0次
            sheet.cell(row=a + 2, column=1).value = int(element)
        for b, element in enumerate(ry1):  # 1万次
            sheet.cell(row=b + 2, column=2).value = int(element)
        for i, element in enumerate(ry2):  # 2万次
            sheet.cell(row=i + 2, column=3).value = int(element)
        for i, element in enumerate(ry3):  # 3万次
            sheet.cell(row=i + 2, column=4).value = int(element)
        for i, element in enumerate(ry4):  # 4万次
            sheet.cell(row=i + 2, column=5).value = int(element)
        for i, element in enumerate(ry5):  # 5万次
            sheet.cell(row=i + 2, column=6).value = int(element)
        for i, element in enumerate(ry6):  # 6万次
            sheet.cell(row=i + 2, column=7).value = int(element)
        for i, element in enumerate(ry7):  # 7万次
            sheet.cell(row=i + 2, column=8).value = int(element)
        wb.save(f'{zuobiao}7万次{yaogan}{guizhong}值.xlsx')

    if zuobiao == 'BL':
        wb = Workbook()
        sheet = wb.active
        sheet.title = '摇杆数据分析'
        fields = [f'{zuobiao}0万次{yaogan}{guizhong}值', f'{zuobiao}1万次{yaogan}{guizhong}值',
                  f'{zuobiao}2万次{yaogan}{guizhong}值', f'{zuobiao}3万次{yaogan}{guizhong}值',
                  f'{zuobiao}4万次{yaogan}{guizhong}值', f'{zuobiao}5万次{yaogan}{guizhong}值',
                  f'{zuobiao}6万次{yaogan}{guizhong}值', f'{zuobiao}7万次{yaogan}{guizhong}值']
        for i in range(len(fields)):  # 首行
            sheet.cell(row=1, column=i + 1).value = fields[i]
        for a, element in enumerate(extra0):  # 0次
            sheet.cell(row=a + 2, column=1).value = int(element)
        for b, element in enumerate(extra1):  # 1万次
            sheet.cell(row=b + 2, column=2).value = int(element)
        for i, element in enumerate(extra2):  # 2万次
            sheet.cell(row=i + 2, column=3).value = int(element)
        for i, element in enumerate(extra3):  # 3万次
            sheet.cell(row=i + 2, column=4).value = int(element)
        for i, element in enumerate(extra4):  # 4万次
            sheet.cell(row=i + 2, column=5).value = int(element)
        for i, element in enumerate(extra5):  # 5万次
            sheet.cell(row=i + 2, column=6).value = int(element)
        for i, element in enumerate(extra6):  # 6万次
            sheet.cell(row=i + 2, column=7).value = int(element)
        for i, element in enumerate(extra7):  # 7万次
            sheet.cell(row=i + 2, column=8).value = int(element)
        wb.save(f'{zuobiao}7万次{yaogan}{guizhong}值.xlsx')


if __name__ == '__main__':
    '''
    归中值
    '''
    # save_data(zuobiao='LY', adc='归中值', yaogan='左摇杆', guizhong='上归中')
    # save_data(zuobiao='LY', adc='归中值', yaogan='左摇杆', guizhong='下归中')
    # save_data(zuobiao='LX', adc='归中值', yaogan='左摇杆', guizhong='左归中')
    # save_data(zuobiao='LX', adc='归中值', yaogan='左摇杆', guizhong='右归中')
    # save_data(zuobiao='RX', adc='归中值', yaogan='右摇杆', guizhong='左归中')
    # save_data(zuobiao='RX', adc='归中值', yaogan='右摇杆', guizhong='右归中')
    # save_data(zuobiao='RY', adc='归中值', yaogan='右摇杆', guizhong='上归中')
    # save_data(zuobiao='RY', adc='归中值', yaogan='右摇杆', guizhong='下归中')
    # save_data(zuobiao='BL', adc='归中值', yaogan='左拨轮', guizhong='左归中')
    # save_data(zuobiao='BL', adc='归中值', yaogan='左拨轮', guizhong='右归中')

    '''
    极限值
    '''
    # save_data(zuobiao='LY', adc='极限值', yaogan='左摇杆', guizhong='上极限值')
    # save_data(zuobiao='LY', adc='极限值', yaogan='左摇杆', guizhong='下极限值')
    # save_data(zuobiao='LX', adc='极限值', yaogan='左摇杆', guizhong='左极限值')
    # save_data(zuobiao='LX', adc='极限值', yaogan='左摇杆', guizhong='右极限值')
    # save_data(zuobiao='RX', adc='极限值', yaogan='右摇杆', guizhong='左极限值')
    # save_data(zuobiao='RX', adc='极限值', yaogan='右摇杆', guizhong='右极限值')
    # save_data(zuobiao='RY', adc='极限值', yaogan='右摇杆', guizhong='上极限值')
    # save_data(zuobiao='RY', adc='极限值', yaogan='右摇杆', guizhong='下极限值')
    # save_data(zuobiao='BL', adc='极限值', yaogan='左拨轮', guizhong='左极限值')
    # save_data(zuobiao='BL', adc='极限值', yaogan='左拨轮', guizhong='右极限值')

    '''
    归中值2
    '''
    # save_data(zuobiao='LX', adc='归中值', yaogan='左摇杆', guizhong='左上归中')
    # save_data(zuobiao='LY', adc='归中值', yaogan='左摇杆', guizhong='左上归中')
    # save_data(zuobiao='LX', adc='归中值', yaogan='左摇杆', guizhong='左下归中')
    # save_data(zuobiao='LY', adc='归中值', yaogan='左摇杆', guizhong='左下归中')
    # save_data(zuobiao='RX', adc='归中值', yaogan='右摇杆', guizhong='右下归中')
    # save_data(zuobiao='RY', adc='归中值', yaogan='右摇杆', guizhong='右下归中')
    # save_data(zuobiao='RX', adc='归中值', yaogan='右摇杆', guizhong='右上归中')
    # save_data(zuobiao='RY', adc='归中值', yaogan='右摇杆', guizhong='右上归中')
    # save_data(zuobiao='LX', adc='归中值', yaogan='左摇杆', guizhong='右上归中')
    # save_data(zuobiao='LY', adc='归中值', yaogan='左摇杆', guizhong='右上归中')
    # save_data(zuobiao='LX', adc='归中值', yaogan='左摇杆', guizhong='右下归中')
    # save_data(zuobiao='LY', adc='归中值', yaogan='左摇杆', guizhong='右下归中')
    # save_data(zuobiao='RX', adc='归中值', yaogan='右摇杆', guizhong='左上归中')
    # save_data(zuobiao='RY', adc='归中值', yaogan='右摇杆', guizhong='左上归中')
    # save_data(zuobiao='RX', adc='归中值', yaogan='右摇杆', guizhong='左下归中')
    # save_data(zuobiao='RY', adc='归中值', yaogan='右摇杆', guizhong='左下归中')

    '''
    极限值2
    '''
    # save_data(zuobiao='LX', adc='极限值', yaogan='左摇杆', guizhong='左上极限值')
    # save_data(zuobiao='LY', adc='极限值', yaogan='左摇杆', guizhong='左上极限值')
    # save_data(zuobiao='LX', adc='极限值', yaogan='左摇杆', guizhong='左下极限值')
    # save_data(zuobiao='LY', adc='极限值', yaogan='左摇杆', guizhong='左下极限值')
    # save_data(zuobiao='LX', adc='极限值', yaogan='左摇杆', guizhong='右上极限值')
    # save_data(zuobiao='LY', adc='极限值', yaogan='左摇杆', guizhong='右上极限值')
    # save_data(zuobiao='LX', adc='极限值', yaogan='左摇杆', guizhong='右下极限值')
    # save_data(zuobiao='LY', adc='极限值', yaogan='左摇杆', guizhong='右下极限值')
    # save_data(zuobiao='RX', adc='极限值', yaogan='右摇杆', guizhong='左上极限值')
    # save_data(zuobiao='RY', adc='极限值', yaogan='右摇杆', guizhong='左上极限值')
    # save_data(zuobiao='RX', adc='极限值', yaogan='右摇杆', guizhong='左下极限值')
    # save_data(zuobiao='RY', adc='极限值', yaogan='右摇杆', guizhong='左下极限值')
    # save_data(zuobiao='RX', adc='极限值', yaogan='右摇杆', guizhong='右上极限值')
    # save_data(zuobiao='RY', adc='极限值', yaogan='右摇杆', guizhong='右上极限值')
    # save_data(zuobiao='RX', adc='极限值', yaogan='右摇杆', guizhong='右下极限值')
    # save_data(zuobiao='RY', adc='极限值', yaogan='右摇杆', guizhong='右下极限值')

    '''
    刻度值
    '''
    # save_data(zuobiao='LX', adc='刻度值', yaogan='左摇杆', guizhong='左右刻度值')
    # save_data(zuobiao='LY', adc='刻度值', yaogan='左摇杆', guizhong='上下刻度值')
    # save_data(zuobiao='RX', adc='刻度值', yaogan='右摇杆', guizhong='左右刻度值')
    # save_data(zuobiao='RY', adc='刻度值', yaogan='右摇杆', guizhong='上下刻度值')
    pass




